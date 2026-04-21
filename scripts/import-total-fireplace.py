#!/usr/bin/env python3
"""
Import Total Fireplace accessories catalog from the XLSX price sheet.

Reads: docs/total-fireplace-2025-pricing.xlsx
Writes: src/data/accessories/total-fireplace.ts

Preserves Wix-CDN product images from the existing generated .ts file by matching
either on exact SKU or via a manual alias table.

Usage:
  python scripts/import-total-fireplace.py
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Optional, TypedDict

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "docs" / "total-fireplace-2025-pricing.xlsx"
OUT_TS = ROOT / "src" / "data" / "accessories" / "total-fireplace.ts"
EXISTING_TS = ROOT / "src" / "data" / "accessories" / "total-fireplace.ts"  # read BEFORE overwriting
INTERMEDIATE_JSON = ROOT / ".scrape" / "tf-parsed.json"

# Aliases: our previously-scraped SKUs (often UPCs or Wix slugs) → XLSX canonical dealer SKUs.
# Used to preserve the Wix-CDN imageUrl across the schema change.
SKU_ALIASES: dict[str, str] = {
    "1001T-LCD-A": "1001 T/LCD",       # Skytech LCD On/Off Remote
    "074197106012": "#601B",           # Meeco Silicone Black 10 oz cartridge
    "074197006138": "#613",            # Meeco Silicone Red cartridge
    "074197001102": "#110",            # Meeco Gasket Cement 3 oz tube
    "LT0162": "LTO162",                # Imperial Black Ash Shovel (source typo)
    "SLRM/LR44": "SLRM",               # Hy-C 45" Log Rack
    "720038002219": "SPEEDY WHITE",    # Speedy White Hearth & Stove Cleaner
}


class Product(TypedDict, total=False):
    category: str
    manufacturer: Optional[str]
    sku: str
    name: str
    variant: Optional[str]
    description: Optional[str]
    wholesale: float
    shipping_tier: str  # "Light" or "Heavy"
    shipping_pct: float  # decimal: 0.08 or 0.16
    retail: int
    image_url: Optional[str]
    source_url: Optional[str]


# ------------------------- Helpers -------------------------

def normalize(text) -> str:
    if text is None:
        return ""
    text = str(text).replace("\ufffd", "®").replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def slugify(text: str) -> str:
    text = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return text[:40] or "item"


def extract_existing_images() -> dict[str, dict[str, str]]:
    """Parse the current total-fireplace.ts and return {sku: {imageUrl, sourceUrl}}."""
    if not EXISTING_TS.exists():
        return {}
    text = EXISTING_TS.read_text(encoding="utf-8")
    images: dict[str, dict[str, str]] = {}
    blocks = re.findall(r"\{[^{}]*\}", text, re.DOTALL)
    for b in blocks:
        sku_m = re.search(r"sku:\s*'([^']+)'", b)
        img_m = re.search(r"imageUrl:\s*'([^']+)'", b)
        src_m = re.search(r"sourceUrl:\s*'([^']+)'", b)
        if sku_m and img_m:
            sku = sku_m.group(1)
            images[sku] = {
                "imageUrl": img_m.group(1),
                "sourceUrl": src_m.group(1) if src_m else "",
            }
    return images


def pick_image_for_sku(xlsx_sku: str, existing_images: dict[str, dict[str, str]]) -> Optional[dict]:
    """Find existing image for this XLSX SKU — try direct match, then aliases."""
    if xlsx_sku in existing_images:
        return existing_images[xlsx_sku]
    # Check if any previous-SKU aliases map to this XLSX SKU
    for old_sku, new_sku in SKU_ALIASES.items():
        if new_sku == xlsx_sku and old_sku in existing_images:
            return existing_images[old_sku]
    return None


# ------------------------- XLSX parser -------------------------

def parse_xlsx(xlsx_path: Path) -> list[Product]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Retail Pricing"]

    existing_images = extract_existing_images()
    print(f"Found {len(existing_images)} existing products with images")

    products: list[Product] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[2]:
            continue
        category = normalize(row[0])
        manufacturer = normalize(row[1])
        sku = normalize(row[2])
        name = normalize(row[3])
        variant = normalize(row[4])
        description = normalize(row[5])
        try:
            wholesale = float(row[6]) if row[6] is not None else 0.0
        except (TypeError, ValueError):
            wholesale = 0.0
        tier = normalize(row[7])
        pct = 0.16 if tier.lower() == "heavy" else 0.08
        try:
            retail = int(row[9]) if row[9] is not None else round(wholesale * (2 + pct))
        except (TypeError, ValueError):
            retail = round(wholesale * (2 + pct))

        img = pick_image_for_sku(sku, existing_images)

        products.append(Product(
            category=category,
            manufacturer=manufacturer,
            sku=sku,
            name=name,
            variant=variant or None,
            description=description or None,
            wholesale=wholesale,
            shipping_tier=tier,
            shipping_pct=pct,
            retail=retail,
            image_url=img["imageUrl"] if img else None,
            source_url=img["sourceUrl"] if img and img.get("sourceUrl") else None,
        ))
    return products


# ------------------------- TypeScript generator -------------------------

def ts_string(s: Optional[str]) -> str:
    if s is None or s == "":
        return "undefined"
    return "'" + s.replace("\\", "\\\\").replace("'", "\\'") + "'"


def product_id(sku: str, name: str, variant: Optional[str]) -> str:
    parts = ["tf", slugify(sku), slugify(name)]
    if variant:
        parts.append(slugify(variant))
    return "-".join(p for p in parts if p)[:60]


def emit_ts(products: list[Product]) -> str:
    # Preserve XLSX row order inside each category; order categories by first-appearance.
    categories: list[str] = []
    seen = set()
    for p in products:
        if p["category"] and p["category"] not in seen:
            categories.append(p["category"])
            seen.add(p["category"])

    lines = [
        "// AUTO-GENERATED — do NOT edit by hand.",
        "// Edit docs/total-fireplace-2025-pricing.xlsx and re-run `npm run import:tf`",
        "// (or `python scripts/import-total-fireplace.py`).",
        "",
        "import { AccessoryVendor } from '@/types/accessories';",
        "",
        "export const totalFireplace: AccessoryVendor = {",
        "  id: 'total-fireplace',",
        "  name: 'Total Fireplace',",
        "  orderEmail: 'info@hibernation.com', // testing — update when live",
        "  categories: [",
    ]
    for c in categories:
        lines.append(f"    {ts_string(c)},")
    lines.append("  ],")
    lines.append("  products: [")

    last_cat = None
    for p in products:
        if p["category"] != last_cat:
            lines.append("")
            lines.append(f"    // ------------------ {p['category']} ------------------")
            last_cat = p["category"]

        pid = product_id(p["sku"], p["name"], p.get("variant"))

        # Build display name: name + (variant) if variant present
        display_name = p["name"]
        if p.get("variant"):
            display_name = f"{p['name']} — {p['variant']}"

        parts = [
            f"id: {ts_string(pid)}",
            f"name: {ts_string(display_name)}",
            f"sku: {ts_string(p['sku'])}",
        ]
        if p.get("manufacturer"):
            parts.append(f"brand: {ts_string(p['manufacturer'])}")
        parts.append(f"category: {ts_string(p['category'])}")
        # `price` is the line-item price used on the order — we pay the vendor wholesale.
        parts.append(f"price: {p['wholesale']:.2f}")
        parts.append(f"wholesale: {p['wholesale']:.2f}")
        parts.append(f"shippingPct: {int(p['shipping_pct'] * 100)}")  # 8 (Light) or 16 (Heavy)
        parts.append(f"storeRetail: {p['retail']}")  # Jeremy's customer-facing price
        if p.get("image_url"):
            parts.append(f"imageUrl: {ts_string(p['image_url'])}")
        if p.get("source_url"):
            parts.append(f"sourceUrl: {ts_string(p['source_url'])}")
        if p.get("description"):
            parts.append(f"description: {ts_string(p['description'])}")

        body = ",\n      ".join(parts)
        lines.append("    {")
        lines.append(f"      {body},")
        lines.append("    },")

    lines.append("  ],")
    lines.append("};")
    lines.append("")
    return "\n".join(lines)


# ------------------------- Main -------------------------

def main():
    if not XLSX_PATH.exists():
        raise SystemExit(f"XLSX not found: {XLSX_PATH}")
    print(f"Parsing {XLSX_PATH.name}...")
    products = parse_xlsx(XLSX_PATH)
    print(f"Parsed {len(products)} products.")

    by_cat: dict[str, int] = {}
    for p in products:
        by_cat[p["category"]] = by_cat.get(p["category"], 0) + 1
    print(f"Categories ({len(by_cat)}):")
    for c, n in sorted(by_cat.items(), key=lambda x: -x[1]):
        print(f"  {c}: {n}")

    with_img = sum(1 for p in products if p.get("image_url"))
    print(f"Products with images: {with_img}/{len(products)}")

    INTERMEDIATE_JSON.parent.mkdir(parents=True, exist_ok=True)
    with INTERMEDIATE_JSON.open("w", encoding="utf-8") as f:
        json.dump(products, f, indent=2, ensure_ascii=False)
    print(f"Wrote {INTERMEDIATE_JSON.relative_to(ROOT)}")

    ts = emit_ts(products)
    OUT_TS.write_text(ts, encoding="utf-8")
    print(f"Wrote {OUT_TS.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
